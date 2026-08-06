import datetime
import random
import io
import base64
import csv
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import qrcode

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, Http404
from django.urls import reverse
from django.utils import timezone

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import (
    CustomUser, Department, Member, Visitor, AttendanceSession, AttendanceRecord,
    Event, FinanceTransaction, Sermon, Announcement, AuditLog, log_activity
)
from .forms import (
    MemberForm, VisitorForm, AttendanceSessionForm, FinanceTransactionForm,
    SermonForm, EventForm, AnnouncementForm, DepartmentForm,
    CustomUserCreationForm, CustomUserUpdateForm, AdminPasswordChangeForm
)
from .decorators import role_required


# --- AUTHENTICATION ---

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        u = request.POST.get('username', '').strip()
        p = request.POST.get('password', '')

        # Check existing user model for lockout state
        user_obj = CustomUser.objects.filter(username__iexact=u).first() if u else None
        if user_obj and user_obj.is_locked_out():
            remaining_seconds = int((user_obj.lockout_until - timezone.now()).total_seconds())
            remaining_mins = max(1, (remaining_seconds + 59) // 60)
            log_activity(request, 'LOGIN', 'Authentication', f"Blocked login attempt for locked-out user '{u}'")
            messages.error(
                request,
                f"Account '{u}' is temporarily locked due to 3 failed login attempts. "
                f"Please try again in {remaining_mins} minute(s) or contact a Super Admin."
            )
            return render(request, 'login.html')

        user = authenticate(request, username=u, password=p)
        if user is not None:
            if user.is_locked_out():
                remaining_seconds = int((user.lockout_until - timezone.now()).total_seconds())
                remaining_mins = max(1, (remaining_seconds + 59) // 60)
                messages.error(
                    request,
                    f"Account '{user.username}' is locked out. Please wait {remaining_mins} minute(s)."
                )
                return render(request, 'login.html')
            
            # Reset lockout tracking on successful login
            user.reset_lockout()
            login(request, user)
            log_activity(request, 'LOGIN', 'Authentication', f"User '{user.username}' ({user.role}) logged in successfully")
            messages.success(request, f"Welcome back, {user.username}!")
            return redirect('dashboard')
        else:
            if user_obj:
                user_obj.register_failed_login()
                if user_obj.is_locked_out():
                    log_activity(request, 'LOGIN', 'Authentication', f"User '{u}' locked out after 3 failed login attempts")
                    messages.error(
                        request,
                        f"Account '{u}' has been locked for 15 minutes after 3 failed login attempts."
                    )
                else:
                    attempts_left = 3 - user_obj.failed_login_attempts
                    log_activity(request, 'LOGIN', 'Authentication', f"Failed login attempt for username '{u}' ({user_obj.failed_login_attempts}/3)")
                    messages.error(
                        request,
                        f"Invalid username or password. {attempts_left} attempt(s) remaining before account lockout."
                    )
            else:
                log_activity(request, 'LOGIN', 'Authentication', f"Failed login attempt for non-existent username '{u}'")
                messages.error(request, "Invalid username or password.")
    return render(request, 'login.html')



@login_required
def logout_view(request):
    if request.user.is_authenticated:
        log_activity(request, 'LOGOUT', 'Authentication', f"User '{request.user.username}' logged out")
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')


# --- DASHBOARD ---

@login_required
def dashboard(request):
    today = datetime.date.today()
    
    # Basic statistics
    total_members = Member.objects.count()
    total_visitors = Visitor.objects.count()
    
    # Today's attendance
    today_sessions = AttendanceSession.objects.filter(date=today)
    today_attendance = AttendanceRecord.objects.filter(session__in=today_sessions, status='Present').count()
    
    # Upcoming events (next 30 days)
    upcoming_events = Event.objects.filter(date__gte=today, date__lte=today + datetime.timedelta(days=30)).order_by('date')
    
    # Total offerings/income this month
    start_of_month = today.replace(day=1)
    monthly_income = FinanceTransaction.objects.filter(
        transaction_type='Income',
        date__gte=start_of_month,
        date__lte=today
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # Recent registrations (last 5)
    recent_registrations = Member.objects.all().order_by('-date_joined')[:5]
    
    # Birthdays in the next 7 days
    birthdays = []
    active_members = Member.objects.filter(status='Active')
    for m in active_members:
        if m.date_of_birth:
            # Check birthday this year
            try:
                b_day = m.date_of_birth.replace(year=today.year)
            except ValueError:  # Handle leap year Feb 29
                b_day = m.date_of_birth.replace(year=today.year, day=28)
            
            if b_day < today:
                # If birthday passed, check next year
                try:
                    b_day = b_day.replace(year=today.year + 1)
                except ValueError:
                    b_day = b_day.replace(year=today.year + 1, day=28)
            
            delta = (b_day - today).days
            if 0 <= delta <= 7:
                birthdays.append({
                    'member': m,
                    'days_until': delta,
                    'date': b_day
                })
    birthdays.sort(key=lambda x: x['days_until'])

    # Weekly Attendance Data for Chart.js
    last_4_sundays = []
    for i in range(4):
        last_4_sundays.append(today - datetime.timedelta(days=today.weekday() + 1 + 7*i))
    last_4_sundays.reverse()
    
    chart_labels = []
    chart_data = []
    for sun_date in last_4_sundays:
        chart_labels.append(sun_date.strftime('%b %d'))
        sun_sess = AttendanceSession.objects.filter(service_type='Sunday Service', date=sun_date).first()
        if sun_sess:
            count = AttendanceRecord.objects.filter(session=sun_sess, status='Present').count()
            chart_data.append(count)
        else:
            chart_data.append(0)

    # Announcements (latest 3)
    announcements = Announcement.objects.all().order_by('-date_posted')[:3]

    # Audit Logs (latest 8)
    recent_audit_logs = AuditLog.objects.all()[:8]

    context = {
        'total_members': total_members,
        'total_visitors': total_visitors,
        'today_attendance': today_attendance,
        'upcoming_events_count': upcoming_events.count(),
        'upcoming_events': upcoming_events[:3],
        'monthly_income': monthly_income,
        'recent_registrations': recent_registrations,
        'birthdays': birthdays,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'announcements': announcements,
        'recent_audit_logs': recent_audit_logs,
    }
    return render(request, 'dashboard.html', context)


# --- MEMBER MANAGEMENT ---

@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def member_list(request):
    query = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    dept_filter = request.GET.get('department', '')
    
    members = Member.objects.all().order_by('full_name')
    
    if query:
        members = members.filter(
            Q(full_name__icontains=query) |
            Q(membership_id__icontains=query) |
            Q(phone_number__icontains=query) |
            Q(cell_group__icontains=query)
        )
    if status_filter:
        members = members.filter(status=status_filter)
    if dept_filter:
        members = members.filter(department_id=dept_filter)
        
    departments = Department.objects.all()
    
    context = {
        'members': members,
        'departments': departments,
        'query': query,
        'status_filter': status_filter,
        'dept_filter': dept_filter,
    }
    return render(request, 'members/member_list.html', context)


@login_required
@role_required(['Super Admin', 'Secretary'])
def member_create(request):
    if request.method == 'POST':
        form = MemberForm(request.POST, request.FILES)
        if form.is_valid():
            member = form.save()
            log_activity(request, 'CREATE', 'Members', f"Registered member '{member.full_name}' (ID: {member.membership_id})")
            messages.success(request, f"Member '{member.full_name}' registered successfully with ID: {member.membership_id}")
            return redirect('member_list')
    else:
        form = MemberForm()
    return render(request, 'members/member_form.html', {'form': form, 'title': 'Register Member'})


@login_required
@role_required(['Super Admin', 'Secretary'])
def member_update(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        form = MemberForm(request.POST, request.FILES, instance=member)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'Members', f"Updated member profile for '{member.full_name}' (ID: {member.membership_id})")
            messages.success(request, f"Member '{member.full_name}' updated successfully.")
            return redirect('member_profile', pk=member.pk)
    else:
        form = MemberForm(instance=member)
    return render(request, 'members/member_form.html', {'form': form, 'title': 'Update Member', 'member': member})


@login_required
@role_required(['Super Admin', 'Pastor'])
def member_delete(request, pk):
    member = get_object_or_404(Member, pk=pk)
    if request.method == 'POST':
        name = member.full_name
        member.delete()
        log_activity(request, 'DELETE', 'Members', f"Deleted member record for '{name}' (ID: {pk})")
        messages.success(request, f"Member '{name}' has been deleted.")
        return redirect('member_list')
    return render(request, 'confirm_delete.html', {'object': member, 'type': 'Member', 'cancel_url': 'member_list'})


@login_required
def member_profile(request, pk):
    member = get_object_or_404(Member, pk=pk)
    # Check permissions (Department leaders can only see members in their department, other staff can see all)
    if request.user.role == 'Department Leader':
        dept_led = Department.objects.filter(leader=request.user).first()
        if not dept_led or member.department != dept_led:
            messages.error(request, "Permission Denied: You can only view members of your department.")
            return redirect('dashboard')
            
    attendance = AttendanceRecord.objects.filter(member=member).order_by('-session__date')
    context = {
        'member': member,
        'attendance': attendance,
    }
    return render(request, 'members/member_profile.html', context)


@login_required
def member_card_print(request, pk):
    member = get_object_or_404(Member, pk=pk)
    return render(request, 'members/member_card_print.html', {'member': member})


# --- VISITOR MANAGEMENT ---

@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def visitor_list(request):
    query = request.GET.get('q', '')
    visitors = Visitor.objects.all().order_by('-first_visit_date')
    if query:
        visitors = visitors.filter(
            Q(name__icontains=query) |
            Q(visitor_id__icontains=query) |
            Q(phone__icontains=query) |
            Q(invited_by__icontains=query)
        )
    return render(request, 'visitors/visitor_list.html', {'visitors': visitors, 'query': query})


@login_required
@role_required(['Super Admin', 'Secretary'])
def visitor_create(request):
    if request.method == 'POST':
        form = VisitorForm(request.POST)
        if form.is_valid():
            v = form.save()
            log_activity(request, 'CREATE', 'Visitors', f"Registered visitor '{v.name}' (ID: {v.visitor_id})")
            messages.success(request, f"Visitor '{v.name}' registered successfully with ID: {v.visitor_id}")
            return redirect('visitor_list')
    else:
        form = VisitorForm()
    return render(request, 'visitors/visitor_form.html', {'form': form, 'title': 'Register Visitor'})


@login_required
@role_required(['Super Admin', 'Secretary'])
def visitor_update(request, pk):
    visitor = get_object_or_404(Visitor, pk=pk)
    if request.method == 'POST':
        form = VisitorForm(request.POST, instance=visitor)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'Visitors', f"Updated visitor profile for '{visitor.name}' (ID: {visitor.visitor_id})")
            messages.success(request, f"Visitor '{visitor.name}' updated successfully.")
            return redirect('visitor_list')
    else:
        form = VisitorForm(instance=visitor)
    return render(request, 'visitors/visitor_form.html', {'form': form, 'title': 'Update Visitor', 'visitor': visitor})


@login_required
@role_required(['Super Admin'])
def visitor_delete(request, pk):
    visitor = get_object_or_404(Visitor, pk=pk)
    if request.method == 'POST':
        name = visitor.name
        visitor.delete()
        log_activity(request, 'DELETE', 'Visitors', f"Deleted visitor record for '{name}' (ID: {pk})")
        messages.success(request, f"Visitor '{name}' has been deleted.")
        return redirect('visitor_list')
    return render(request, 'confirm_delete.html', {'object': visitor, 'type': 'Visitor', 'cancel_url': 'visitor_list'})


# --- ATTENDANCE MANAGEMENT ---

@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary', 'Department Leader'])
def attendance_list(request):
    sessions = AttendanceSession.objects.all().order_by('-date')
    return render(request, 'attendance/attendance_list.html', {'sessions': sessions})


@login_required
@role_required(['Super Admin', 'Secretary'])
def attendance_session_create(request):
    if request.method == 'POST':
        form = AttendanceSessionForm(request.POST)
        if form.is_valid():
            session = form.save()
            # Prefill all active members as Present for this session by default
            active_members = Member.objects.filter(status='Active')
            for m in active_members:
                AttendanceRecord.objects.create(session=session, member=m, status='Absent')
            log_activity(request, 'CREATE', 'Attendance', f"Created attendance session '{session.service_type}' for {session.date}")
            messages.success(request, f"Attendance session for {session.service_type} on {session.date} created. Please mark attendance.")
            return redirect('attendance_session_detail', pk=session.pk)
    else:
        form = AttendanceSessionForm(initial={'date': datetime.date.today()})
    return render(request, 'attendance/session_form.html', {'form': form, 'title': 'Create Attendance Session'})


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary', 'Department Leader'])
def attendance_session_detail(request, pk):
    session = get_object_or_404(AttendanceSession, pk=pk)
    
    # Save attendance updates
    if request.method == 'POST':
        if request.user.role not in ['Super Admin', 'Secretary']:
            messages.error(request, "Permission Denied: Only Secretary or Super Admin can record attendance.")
            return redirect('attendance_session_detail', pk=session.pk)
            
        present_member_ids = request.POST.getlist('members_present')
        present_visitor_ids = request.POST.getlist('visitors_present')
        
        # Update Member records for this session
        all_records = AttendanceRecord.objects.filter(session=session)
        for rec in all_records:
            if rec.member:
                rec.status = 'Present' if str(rec.member.pk) in present_member_ids else 'Absent'
                rec.save()
            elif rec.visitor:
                rec.status = 'Present' if str(rec.visitor.pk) in present_visitor_ids else 'Absent'
                rec.save()
                
        # Handle adding visitors that were present but not previously in this session
        for vis_id in present_visitor_ids:
            if not all_records.filter(visitor_id=vis_id).exists():
                v = Visitor.objects.filter(pk=vis_id).first()
                if v:
                    AttendanceRecord.objects.create(session=session, visitor=v, status='Present')
                    
        log_activity(request, 'UPDATE', 'Attendance', f"Recorded attendance for '{session.service_type}' on {session.date}")
        messages.success(request, "Attendance logged successfully.")
        return redirect('attendance_session_detail', pk=session.pk)

    # Restrict viewing department members if the user is a Department Leader
    dept_led = None
    if request.user.role == 'Department Leader':
        dept_led = Department.objects.filter(leader=request.user).first()

    member_records = AttendanceRecord.objects.filter(session=session, member__isnull=False).select_related('member')
    if dept_led:
        member_records = member_records.filter(member__department=dept_led)
        
    visitor_records = AttendanceRecord.objects.filter(session=session, visitor__isnull=False).select_related('visitor')
    
    # Visitors available to add (registered on or before service date)
    available_visitors = Visitor.objects.filter(first_visit_date__lte=session.date)
    
    total_present = AttendanceRecord.objects.filter(session=session, status='Present').count()
    total_absent = AttendanceRecord.objects.filter(session=session, status='Absent').count()
    
    context = {
        'session': session,
        'member_records': member_records,
        'visitor_records': visitor_records,
        'available_visitors': available_visitors,
        'total_present': total_present,
        'total_absent': total_absent,
        'dept_led': dept_led,
    }
    return render(request, 'attendance/session_detail.html', context)


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary', 'Department Leader'])
def attendance_session_qr(request, pk):
    session = get_object_or_404(AttendanceSession, pk=pk)
    checkin_url = request.build_absolute_uri(reverse('attendance_self_checkin', kwargs={'pk': session.pk}))
    
    # Generate QR Code image in base64
    qr = qrcode.QRCode(version=1, box_size=10, border=3)
    qr.add_data(checkin_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    qr_b64 = base64.b64encode(buffer.getvalue()).decode('utf-8')

    context = {
        'session': session,
        'checkin_url': checkin_url,
        'qr_b64': qr_b64,
    }
    return render(request, 'attendance/session_qr.html', context)


def attendance_self_checkin(request, pk):
    session = get_object_or_404(AttendanceSession, pk=pk)
    success = False
    message = None
    error = None
    member = None
    visitor = None

    if request.method == 'POST':
        user_type = request.POST.get('user_type', 'member')
        
        if user_type == 'member':
            identifier = request.POST.get('identifier', '').strip()
            if not identifier:
                error = "Please enter your Membership ID or Phone Number."
            else:
                matched_member = Member.objects.filter(
                    Q(membership_id__iexact=identifier) |
                    Q(phone_number__iexact=identifier) |
                    Q(full_name__iexact=identifier)
                ).first()

                if not matched_member:
                    matched_member = Member.objects.filter(
                        Q(phone_number__icontains=identifier) |
                        Q(full_name__icontains=identifier)
                    ).first()

                if matched_member:
                    member = matched_member
                    rec, created = AttendanceRecord.objects.get_or_create(
                        session=session,
                        member=member,
                        defaults={'status': 'Present'}
                    )
                    if not created:
                        rec.status = 'Present'
                        rec.save()
                    success = True
                    message = f"Welcome, {member.full_name}! Your attendance for {session.service_type} on {session.date} has been recorded as PRESENT."
                    log_activity(request, 'CREATE', 'Attendance', f"Self check-in completed for member '{member.full_name}' in '{session.service_type}'")
                else:
                    error = f"No member found matching '{identifier}'. Please check your Membership ID (e.g. RACI26/001) or register as a guest."
        
        elif user_type == 'visitor':
            name = request.POST.get('visitor_name', '').strip()
            phone = request.POST.get('visitor_phone', '').strip()
            address = request.POST.get('visitor_address', '').strip() or 'Guest Check-in'

            if not name or not phone:
                error = "Please provide both your Name and Phone Number to check in as a guest."
            else:
                vis, _ = Visitor.objects.get_or_create(
                    phone=phone,
                    defaults={'name': name, 'address': address, 'first_visit_date': session.date}
                )
                rec, created = AttendanceRecord.objects.get_or_create(
                    session=session,
                    visitor=vis,
                    defaults={'status': 'Present'}
                )
                if not created:
                    rec.status = 'Present'
                    rec.save()
                visitor = vis
                success = True
                message = f"Welcome, {vis.name}! Your attendance as a guest for {session.service_type} has been recorded as PRESENT."
                log_activity(request, 'CREATE', 'Attendance', f"Self check-in completed for guest '{vis.name}' in '{session.service_type}'")

    all_members = Member.objects.filter(status='Active').order_by('full_name')

    context = {
        'session': session,
        'success': success,
        'message': message,
        'error': error,
        'member': member,
        'visitor': visitor,
        'all_members': all_members,
    }
    return render(request, 'attendance/self_checkin.html', context)


# --- DEPARTMENTS ---

@login_required
def department_list(request):
    departments = Department.objects.annotate(member_count=Count('members'))
    return render(request, 'departments/department_list.html', {'departments': departments})


@login_required
@role_required(['Super Admin', 'Pastor'])
def department_create(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            dept = form.save()
            log_activity(request, 'CREATE', 'Departments', f"Created department '{dept.name}'")
            messages.success(request, "Department created successfully.")
            return redirect('department_list')
    else:
        form = DepartmentForm()
    return render(request, 'departments/department_form.html', {'form': form, 'title': 'Add Department'})


@login_required
@role_required(['Super Admin', 'Pastor'])
def department_update(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=dept)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'Departments', f"Updated department '{dept.name}'")
            messages.success(request, f"Department '{dept.name}' updated successfully.")
            return redirect('department_list')
    else:
        form = DepartmentForm(instance=dept)
    return render(request, 'departments/department_form.html', {'form': form, 'title': 'Edit Department'})


@login_required
def department_detail(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    
    # Role check: Department Leaders can only see their own department details
    if request.user.role == 'Department Leader':
        led_dept = Department.objects.filter(leader=request.user).first()
        if not led_dept or led_dept != dept:
            messages.error(request, "Permission Denied: You can only view your own department.")
            return redirect('dashboard')
            
    members = dept.members.all()
    context = {
        'department': dept,
        'members': members,
    }
    return render(request, 'departments/department_detail.html', context)


# --- FINANCE MODULE ---

@login_required
@role_required(['Super Admin', 'Finance Officer'])
def finance_dashboard(request):
    today = datetime.date.today()
    transactions = FinanceTransaction.objects.all().order_by('-date')
    
    # Calculate Income vs Expenses
    total_income = FinanceTransaction.objects.filter(transaction_type='Income').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_expense = FinanceTransaction.objects.filter(transaction_type='Expense').aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    net_balance = total_income - total_expense
    
    # Category Breakdowns
    category_summary = FinanceTransaction.objects.values('category', 'transaction_type').annotate(total=Sum('amount'))
    
    income_categories = []
    income_amounts = []
    expense_categories = []
    expense_amounts = []
    
    for item in category_summary:
        if item['transaction_type'] == 'Income':
            income_categories.append(item['category'])
            income_amounts.append(float(item['total']))
        else:
            expense_categories.append(item['category'])
            expense_amounts.append(float(item['total']))
            
    context = {
        'transactions': transactions[:10],
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'income_categories': income_categories,
        'income_amounts': income_amounts,
        'expense_categories': expense_categories,
        'expense_amounts': expense_amounts,
    }
    return render(request, 'finance/finance_dashboard.html', context)


@login_required
@role_required(['Super Admin', 'Finance Officer'])
def finance_transaction_create(request):
    if request.method == 'POST':
        form = FinanceTransactionForm(request.POST)
        if form.is_valid():
            tx = form.save(commit=False)
            tx.recorded_by = request.user
            tx.save()
            log_activity(request, 'CREATE', 'Finance', f"Logged {tx.transaction_type} of GHS {tx.amount:.2f} ({tx.category})")
            messages.success(request, f"{tx.transaction_type} transaction of GHS {tx.amount} logged successfully.")
            return redirect('finance_dashboard')
    else:
        form = FinanceTransactionForm(initial={'date': datetime.date.today()})
    return render(request, 'finance/transaction_form.html', {'form': form, 'title': 'Log Transaction'})


@login_required
@role_required(['Super Admin', 'Finance Officer'])
def finance_transaction_update(request, pk):
    tx = get_object_or_404(FinanceTransaction, pk=pk)
    if request.method == 'POST':
        form = FinanceTransactionForm(request.POST, instance=tx)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'Finance', f"Updated transaction #{tx.pk} ({tx.transaction_type} - GHS {tx.amount:.2f})")
            messages.success(request, "Transaction updated successfully.")
            return redirect('finance_dashboard')
    else:
        form = FinanceTransactionForm(instance=tx)
    return render(request, 'finance/transaction_form.html', {'form': form, 'title': 'Update Transaction', 'transaction': tx})


@login_required
@role_required(['Super Admin'])
def finance_transaction_delete(request, pk):
    tx = get_object_or_404(FinanceTransaction, pk=pk)
    if request.method == 'POST':
        amount = tx.amount
        category = tx.category
        tx.delete()
        log_activity(request, 'DELETE', 'Finance', f"Deleted transaction '{category} : GHS {amount:.2f}'")
        messages.success(request, f"Transaction '{category} : {amount}' deleted.")
        return redirect('finance_dashboard')
    return render(request, 'confirm_delete.html', {'object': tx, 'type': 'Transaction', 'cancel_url': 'finance_dashboard'})


# --- SERMON MANAGEMENT ---

@login_required
def sermon_list(request):
    query = request.GET.get('q', '')
    sermons = Sermon.objects.all().order_by('-date')
    if query:
        sermons = sermons.filter(
            Q(title__icontains=query) |
            Q(speaker__icontains=query) |
            Q(scripture__icontains=query) |
            Q(notes__icontains=query)
        )
    return render(request, 'sermons/sermon_list.html', {'sermons': sermons, 'query': query})


@login_required
@role_required(['Super Admin', 'Pastor'])
def sermon_create(request):
    if request.method == 'POST':
        form = SermonForm(request.POST, request.FILES)
        if form.is_valid():
            s = form.save()
            log_activity(request, 'CREATE', 'Sermons', f"Added sermon '{s.title}' by {s.speaker}")
            messages.success(request, f"Sermon '{s.title}' added successfully.")
            return redirect('sermon_list')
    else:
        form = SermonForm(initial={'date': datetime.date.today()})
    return render(request, 'sermons/sermon_form.html', {'form': form, 'title': 'Add Sermon'})


@login_required
@role_required(['Super Admin', 'Pastor'])
def sermon_update(request, pk):
    sermon = get_object_or_404(Sermon, pk=pk)
    if request.method == 'POST':
        form = SermonForm(request.POST, request.FILES, instance=sermon)
        if form.is_valid():
            s = form.save()
            log_activity(request, 'UPDATE', 'Sermons', f"Updated sermon '{s.title}'")
            messages.success(request, "Sermon details updated successfully.")
            return redirect('sermon_list')
    else:
        form = SermonForm(instance=sermon)
    return render(request, 'sermons/sermon_form.html', {'form': form, 'title': 'Edit Sermon', 'sermon': sermon})


@login_required
@role_required(['Super Admin'])
def sermon_delete(request, pk):
    sermon = get_object_or_404(Sermon, pk=pk)
    if request.method == 'POST':
        title = sermon.title
        sermon.delete()
        log_activity(request, 'DELETE', 'Sermons', f"Deleted sermon '{title}'")
        messages.success(request, f"Sermon '{title}' deleted.")
        return redirect('sermon_list')
    return render(request, 'confirm_delete.html', {'object': sermon, 'type': 'Sermon', 'cancel_url': 'sermon_list'})


# --- EVENTS ---

@login_required
def event_list(request):
    events = Event.objects.all().order_by('-date')
    return render(request, 'events/event_list.html', {'events': events})


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def event_create(request):
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            ev = form.save()
            log_activity(request, 'CREATE', 'Events', f"Created event '{ev.event_name}' scheduled for {ev.date}")
            messages.success(request, f"Event '{ev.event_name}' created successfully.")
            return redirect('event_list')
    else:
        form = EventForm()
    return render(request, 'events/event_form.html', {'form': form, 'title': 'Create Event'})


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def event_update(request, pk):
    ev = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        form = EventForm(request.POST, instance=ev)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'Events', f"Updated event '{ev.event_name}'")
            messages.success(request, f"Event '{ev.event_name}' updated successfully.")
            return redirect('event_list')
    else:
        form = EventForm(instance=ev)
    return render(request, 'events/event_form.html', {'form': form, 'title': 'Update Event', 'event': ev})


@login_required
@role_required(['Super Admin'])
def event_delete(request, pk):
    ev = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        name = ev.event_name
        ev.delete()
        log_activity(request, 'DELETE', 'Events', f"Deleted event '{name}'")
        messages.success(request, f"Event '{name}' deleted successfully.")
        return redirect('event_list')
    return render(request, 'confirm_delete.html', {'object': ev, 'type': 'Event', 'cancel_url': 'event_list'})


# --- ANNOUNCEMENTS ---

@login_required
def announcement_list(request):
    announcements = Announcement.objects.all().order_by('-date_posted')
    return render(request, 'announcements/announcement_list.html', {'announcements': announcements})


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def announcement_create(request):
    if request.method == 'POST':
        form = AnnouncementForm(request.POST)
        if form.is_valid():
            ann = form.save(commit=False)
            ann.author = request.user
            ann.save()
            log_activity(request, 'CREATE', 'Announcements', f"Published announcement '{ann.title}'")
            messages.success(request, "Announcement published.")
            return redirect('dashboard')
    else:
        form = AnnouncementForm()
    return render(request, 'announcements/announcement_form.html', {'form': form, 'title': 'Publish Announcement'})


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def announcement_update(request, pk):
    ann = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        form = AnnouncementForm(request.POST, instance=ann)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'Announcements', f"Updated announcement '{ann.title}'")
            messages.success(request, "Announcement updated.")
            return redirect('announcement_list')
    else:
        form = AnnouncementForm(instance=ann)
    return render(request, 'announcements/announcement_form.html', {'form': form, 'title': 'Update Announcement', 'announcement': ann})


@login_required
@role_required(['Super Admin', 'Pastor'])
def announcement_delete(request, pk):
    ann = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        title = ann.title
        ann.delete()
        log_activity(request, 'DELETE', 'Announcements', f"Deleted announcement '{title}'")
        messages.success(request, f"Announcement '{title}' deleted.")
        return redirect('announcement_list')
    return render(request, 'confirm_delete.html', {'object': ann, 'type': 'Announcement', 'cancel_url': 'announcement_list'})


# --- DOCUMENT & DATA EXPORT (EXCEL & PDF) ---

@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def export_members_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members Directory"

    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="312E81", end_color="312E81", fill_type="solid")  # Indigo-900 style
    align_center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Membership ID", "Full Name", "Gender", "Date of Birth", 
        "Phone Number", "Address", "Marital Status", "Baptized", 
        "Date Joined", "Department", "Cell Group", "Status"
    ]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    members = Member.objects.all().order_by('full_name')
    for m in members:
        dept = m.department.name if m.department else "None"
        ws.append([
            m.membership_id,
            m.full_name,
            m.gender,
            m.date_of_birth.strftime('%Y-%m-%d') if m.date_of_birth else '',
            m.phone_number,
            m.address,
            m.marital_status,
            "Yes" if m.baptized else "No",
            m.date_joined.strftime('%Y-%m-%d') if m.date_joined else '',
            dept,
            m.cell_group,
            m.status
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="church_members_list.xlsx"'
    wb.save(response)
    log_activity(request, 'EXPORT', 'Members', "Exported member directory to Excel (.xlsx)")
    return response


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def export_members_pdf(request):
    members = Member.objects.all().order_by('full_name')
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#312E81'),
        spaceAfter=15,
        alignment=1
    )

    story.append(Paragraph("Church Members Directory", title_style))
    story.append(Spacer(1, 10))

    data = [["ID", "Full Name", "Gender", "Phone", "Status", "Joined", "Department"]]
    for m in members:
        dept = m.department.name if m.department else "None"
        data.append([
            m.membership_id,
            m.full_name,
            m.gender,
            m.phone_number,
            m.status,
            m.date_joined.strftime('%Y-%m-%d') if m.date_joined else '',
            dept
        ])

    t = Table(data, colWidths=[85, 120, 50, 80, 50, 75, 110])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#312E81')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
    ]))

    story.append(t)
    doc.build(story)
    buffer.seek(0)
    log_activity(request, 'EXPORT', 'Members', "Exported member directory to PDF (.pdf)")
    return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="members_directory.pdf"'})


@login_required
@role_required(['Super Admin', 'Pastor', 'Secretary'])
def download_member_card_pdf(request, pk):
    member = get_object_or_404(Member, pk=pk)
    buffer = io.BytesIO()
    
    # 350x220 points is approximately credit card sized layout
    p = canvas.Canvas(buffer, pagesize=(350, 220))
    
    # Card Background - Deep Royal Purple
    p.setFillColor(colors.HexColor('#1E0A38'))
    p.rect(0, 0, 350, 220, fill=True, stroke=False)
    
    # Top bar - Metallic Gold accent
    p.setFillColor(colors.HexColor('#D4AF37'))
    p.rect(0, 180, 350, 40, fill=True, stroke=False)

    # Gold Border outline
    p.setStrokeColor(colors.HexColor('#F7D070'))
    p.setLineWidth(1.5)
    p.rect(0, 0, 350, 220, fill=False, stroke=True)
    
    # Text Header & Logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')
    text_x = 15
    if os.path.exists(logo_path):
        try:
            p.drawImage(logo_path, 10, 183, width=32, height=34, mask='auto', preserveAspectRatio=True)
            text_x = 48
        except Exception:
            pass

    p.setFillColor(colors.HexColor('#1E0A38'))
    p.setFont("Helvetica-Bold", 12)
    p.drawString(text_x, 193, "R.A.C.I CHURCH")
    
    p.setFont("Helvetica-Bold", 8)
    p.setFillColor(colors.HexColor('#1E0A38'))
    p.drawString(250, 196, "MEMBER CARD")

    # Member Photo
    photo_drawn = False
    if member.photo:
        try:
            photo_path = member.photo.path
            p.drawImage(photo_path, 15, 45, width=90, height=110)
            photo_drawn = True
        except Exception:
            pass

    if not photo_drawn:
        # Placeholder picture
        p.setStrokeColor(colors.HexColor('#D4AF37'))
        p.setFillColor(colors.HexColor('#2E1065'))
        p.rect(15, 45, 90, 110, fill=True, stroke=True)
        p.setFillColor(colors.HexColor('#F7D070'))
        p.setFont("Helvetica-Bold", 10)
        p.drawCentredString(60, 95, "NO PHOTO")

    # Details
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 12)
    p.drawString(120, 140, member.full_name[:22].upper())
    
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor('#F7D070'))
    p.drawString(120, 118, "ID:")
    p.drawString(120, 98, "Dept:")
    p.drawString(120, 78, "Joined:")
    p.drawString(120, 58, "Status:")

    p.setFont("Helvetica", 9)
    p.setFillColor(colors.HexColor('#F3E8FF'))
    p.drawString(140, 118, f"{member.membership_id}")
    p.drawString(155, 98, f"{member.department.name if member.department else 'General'}")
    p.drawString(165, 78, f"{member.date_joined.strftime('%b %d, %Y') if member.date_joined else 'N/A'}")
    p.drawString(160, 58, f"{member.status}")
    
    # Divider line
    p.setStrokeColor(colors.HexColor('#D4AF37'))
    p.setLineWidth(1)
    p.line(120, 48, 335, 48)
    
    # Footer Notice
    p.setFont("Helvetica-Oblique", 7)
    p.setFillColor(colors.HexColor('#E9D5FF'))
    p.drawString(120, 36, "Authorized Member ID. Non-transferable.")
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    log_activity(request, 'EXPORT', 'Members', f"Downloaded digital member ID card for '{member.full_name}' (ID: {member.membership_id})")
    return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': f'attachment; filename="member_card_{member.membership_id}.pdf"'})


@login_required
@role_required(['Super Admin', 'Finance Officer'])
def export_finance_report(request):
    format_type = request.GET.get('format', 'excel')
    transactions = FinanceTransaction.objects.all().order_by('-date')

    if format_type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Transactions"

        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="065F46", end_color="065F46", fill_type="solid")  # Greenish-900 style
        align_center = Alignment(horizontal="center", vertical="center")

        headers = ["Date", "Type", "Category", "Amount (GHS)", "Description", "Recorded By"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        for tx in transactions:
            ws.append([
                tx.date.strftime('%Y-%m-%d'),
                tx.transaction_type,
                tx.category,
                float(tx.amount),
                tx.description,
                tx.recorded_by.username if tx.recorded_by else 'System'
            ])

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="financial_transactions.xlsx"'
        wb.save(response)
        log_activity(request, 'EXPORT', 'Finance', "Exported financial statement report in EXCEL format")
        return response

    elif format_type == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'FinanceDocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=20,
            textColor=colors.HexColor('#065F46'),
            spaceAfter=15,
            alignment=1
        )

        story.append(Paragraph("Church Financial Statement", title_style))
        story.append(Spacer(1, 10))

        data = [["Date", "Type", "Category", "Amount (GHS)", "Recorded By"]]
        total_income = Decimal('0.00')
        total_expense = Decimal('0.00')

        for tx in transactions:
            data.append([
                tx.date.strftime('%Y-%m-%d'),
                tx.transaction_type,
                tx.category,
                f"{tx.amount:.2f}",
                tx.recorded_by.username if tx.recorded_by else 'System'
            ])
            if tx.transaction_type == 'Income':
                total_income += tx.amount
            else:
                total_expense += tx.amount

        # Summary box
        summary_text = f"<b>Total Income:</b> GHS {total_income:.2f} &nbsp;&nbsp;&nbsp;&nbsp; <b>Total Expense:</b> GHS {total_expense:.2f} &nbsp;&nbsp;&nbsp;&nbsp; <b>Net Balance:</b> GHS {(total_income - total_expense):.2f}"
        summary_style = ParagraphStyle('SummaryStyle', parent=styles['Normal'], fontSize=11, spaceAfter=15)
        story.append(Paragraph(summary_text, summary_style))

        t = Table(data, colWidths=[100, 80, 120, 100, 100])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#065F46')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0FDF4')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#A7F3D0')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')])
        ]))

        story.append(t)
        doc.build(story)
        buffer.seek(0)
        log_activity(request, 'EXPORT', 'Finance', "Exported financial statement report in PDF format")
        return HttpResponse(buffer, content_type='application/pdf', headers={'Content-Disposition': 'attachment; filename="financial_statement.pdf"'})

    raise Http404("Invalid format requested.")


# --- USER MANAGEMENT ---

@login_required
@role_required(['Super Admin'])
def user_list(request):
    users = CustomUser.objects.all().order_by('username')
    return render(request, 'users/user_list.html', {'users': users})


@login_required
@role_required(['Super Admin'])
def user_create(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            log_activity(request, 'CREATE', 'User Management', f"Created user account '{user.username}' (Role: {user.role})")
            messages.success(request, f"User '{user.username}' created successfully!")
            return redirect('user_list')
    else:
        form = CustomUserCreationForm()
    return render(request, 'users/user_form.html', {'form': form, 'title': 'Add New User'})


@login_required
@role_required(['Super Admin'])
def user_update(request, pk):
    target_user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        form = CustomUserUpdateForm(request.POST, instance=target_user)
        if form.is_valid():
            form.save()
            log_activity(request, 'UPDATE', 'User Management', f"Updated user profile '{target_user.username}' (Role: {target_user.role})")
            messages.success(request, f"User '{target_user.username}' updated successfully.")
            return redirect('user_list')
    else:
        form = CustomUserUpdateForm(instance=target_user)
    return render(request, 'users/user_form.html', {
        'form': form,
        'title': f'Edit User: {target_user.username}',
        'target_user': target_user
    })


@login_required
@role_required(['Super Admin'])
def user_change_password(request, pk):
    target_user = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        form = AdminPasswordChangeForm(request.POST)
        if form.is_valid():
            new_pass = form.cleaned_data['new_password']
            target_user.set_password(new_pass)
            target_user.save()
            log_activity(request, 'SECURITY', 'User Management', f"Changed password for user account '{target_user.username}'")
            messages.success(request, f"Password for '{target_user.username}' changed successfully.")
            return redirect('user_list')
    else:
        form = AdminPasswordChangeForm()
    return render(request, 'users/change_password.html', {
        'form': form,
        'target_user': target_user
    })


@login_required
@role_required(['Super Admin'])
def user_delete(request, pk):
    target_user = get_object_or_404(CustomUser, pk=pk)
    if target_user == request.user:
        messages.error(request, "You cannot delete your own admin account!")
        return redirect('user_list')
    
    if request.method == 'POST':
        username = target_user.username
        target_user.delete()
        log_activity(request, 'DELETE', 'User Management', f"Deleted user account '{username}' (ID: {pk})")
        messages.success(request, f"User '{username}' has been deleted.")
        return redirect('user_list')
    
    return render(request, 'confirm_delete.html', {
        'object': target_user,
        'type': 'User',
        'cancel_url': 'user_list'
    })


@login_required
@role_required(['Super Admin', 'Pastor'])
def audit_log_list(request):
    query = request.GET.get('q', '').strip()
    module_filter = request.GET.get('module', '').strip()
    action_filter = request.GET.get('action', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    logs = AuditLog.objects.all().select_related('user')

    if query:
        logs = logs.filter(
            Q(description__icontains=query) |
            Q(user_display__icontains=query) |
            Q(ip_address__icontains=query) |
            Q(module__icontains=query)
        )

    if module_filter:
        logs = logs.filter(module=module_filter)

    if action_filter:
        logs = logs.filter(action_type=action_filter)

    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)

    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    total_count = logs.count()
    today = timezone.now().date()
    today_logs_count = logs.filter(timestamp__date=today).count()
    security_alerts_count = logs.filter(
        Q(action_type__in=['SECURITY', 'ACCESS_DENIED']) | 
        Q(description__icontains='lock') | 
        Q(description__icontains='blocked')
    ).count()
    failed_logins_count = logs.filter(
        Q(description__icontains='failed') | 
        Q(description__icontains='blocked') | 
        Q(description__icontains='locked')
    ).count()
    exports_count = logs.filter(action_type='EXPORT').count()

    all_modules = AuditLog.objects.values_list('module', flat=True).distinct()
    all_actions = AuditLog.objects.values_list('action_type', flat=True).distinct()

    context = {
        'logs': logs[:300],
        'query': query,
        'module_filter': module_filter,
        'action_filter': action_filter,
        'date_from': date_from,
        'date_to': date_to,
        'modules': sorted([m for m in set(all_modules) if m]),
        'actions': sorted([a for a in set(all_actions) if a]),
        'total_count': total_count,
        'today_logs_count': today_logs_count,
        'security_alerts_count': security_alerts_count,
        'failed_logins_count': failed_logins_count,
        'exports_count': exports_count,
    }
    return render(request, 'audit_logs/audit_log_list.html', context)


@login_required
@role_required(['Super Admin', 'Pastor'])
def export_audit_logs_csv(request):
    query = request.GET.get('q', '').strip()
    module_filter = request.GET.get('module', '').strip()
    action_filter = request.GET.get('action', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to = request.GET.get('date_to', '').strip()

    logs = AuditLog.objects.all().select_related('user')

    if query:
        logs = logs.filter(
            Q(description__icontains=query) |
            Q(user_display__icontains=query) |
            Q(ip_address__icontains=query) |
            Q(module__icontains=query)
        )

    if module_filter:
        logs = logs.filter(module=module_filter)

    if action_filter:
        logs = logs.filter(action_type=action_filter)

    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)

    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="system_audit_logs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Log ID', 'Timestamp', 'User', 'User Role', 'Action Type', 'Module', 'Description', 'IP Address'])

    for log in logs:
        user_role = log.user.role if log.user else 'System/Guest'
        writer.writerow([
            log.id,
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.user_display,
            user_role,
            log.action_type,
            log.module,
            log.description,
            log.ip_address or 'N/A'
        ])

    log_activity(request, 'EXPORT', 'Audit Logs', f"Exported {logs.count()} activity audit log records to CSV format")
    return response


# --- CONTACT & EMAIL VERIFICATION ---

@login_required
def send_verification_code(request, entity_type, pk, verify_type):
    """
    Generates and sets a 6-digit OTP code for Email or Phone verification.
    entity_type: 'user' or 'member'
    verify_type: 'email' or 'phone'
    """
    code = f"{random.randint(100000, 999999)}"
    if entity_type == 'user':
        target = get_object_or_404(CustomUser, pk=pk)
        target_name = target.username
        contact_val = target.email if verify_type == 'email' else target.phone_number
    elif entity_type == 'member':
        target = get_object_or_404(Member, pk=pk)
        target_name = target.full_name
        contact_val = target.email if verify_type == 'email' else target.phone_number
    else:
        messages.error(request, "Invalid verification target.")
        return redirect('dashboard')

    if not contact_val:
        messages.error(request, f"Cannot verify {verify_type}: No {verify_type} contact detail registered.")
        return redirect('user_list' if entity_type == 'user' else 'member_list')

    if verify_type == 'email':
        target.email_verification_code = code
    else:
        target.phone_verification_code = code
    target.save()

    log_activity(request, 'UPDATE', 'Verification', f"Generated {verify_type} verification code for {entity_type} '{target_name}'")
    messages.info(request, f"Verification code generated for {target_name}'s {verify_type} ({contact_val}). Verification OTP: {code}")
    
    return render(request, 'verification.html', {
        'entity_type': entity_type,
        'pk': pk,
        'verify_type': verify_type,
        'target': target,
        'target_name': target_name,
        'contact_val': contact_val,
        'demo_code': code,
    })


@login_required
def verify_code(request, entity_type, pk, verify_type):
    """
    Validates submitted 6-digit OTP verification code for Email or Phone.
    """
    if entity_type == 'user':
        target = get_object_or_404(CustomUser, pk=pk)
        redirect_url = 'user_list'
        target_name = target.username
    elif entity_type == 'member':
        target = get_object_or_404(Member, pk=pk)
        redirect_url = 'member_list'
        target_name = target.full_name
    else:
        messages.error(request, "Invalid verification target.")
        return redirect('dashboard')

    if request.method == 'POST':
        submitted_code = request.POST.get('code', '').strip()
        expected_code = target.email_verification_code if verify_type == 'email' else target.phone_verification_code

        if expected_code and submitted_code == expected_code:
            if verify_type == 'email':
                target.is_email_verified = True
                target.email_verification_code = None
            else:
                target.is_phone_verified = True
                target.phone_verification_code = None
            target.save()
            log_activity(request, 'UPDATE', 'Verification', f"Successfully verified {verify_type} for {entity_type} '{target_name}'")
            messages.success(request, f"Successfully verified {verify_type} for {target_name}!")
            return redirect(redirect_url)
        else:
            messages.error(request, "Invalid verification code. Please check the code and try again.")

    return redirect('send_verification_code', entity_type=entity_type, pk=pk, verify_type=verify_type)



